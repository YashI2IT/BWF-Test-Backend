"""
Parse all BWF Excel files and dump structured JSON summaries.
"""
import openpyxl, json, os

OUTPUT = "BWF-data/parsed_data.json"
FILES = {
    "students": "BWF-data/Stu.name list jkbwf.xlsx",
    "anantnag": "BWF-data/01 Anantnag  Home Basic requirements.xlsx",
    "kupwara":  "BWF-data/02 Kupwara Home Basic requirements Monthly and Yearly.xlsx",
    "jammu":    "BWF-data/03 Jammu Home Basic requirements Daily_monthly.xlsx",
}

def read_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(c is not None for c in row):
                rows.append([str(c) if c is not None else None for c in row])
        result[sname] = rows
    return result

all_data = {}
for key, path in FILES.items():
    if os.path.exists(path):
        all_data[key] = read_sheet(path)
        print(f"[OK] {key}: {list(all_data[key].keys())}")
    else:
        print(f"[MISSING] {path}")

with open(OUTPUT, "w", encoding="utf-8") as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2, default=str)

print(f"\nSaved to {OUTPUT}")
